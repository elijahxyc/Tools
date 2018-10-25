import xlrd
import xlwt
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


readfile_list = [] #文件列表
MAX = 3 #文件的数目；时间
date_start = 20180820
para = ["investment_fund","security_type","security_category","security_name","coupon_rate","face_value","duration","weight","issue_type","issue_country","date"]
all_data = [] #汇总的数据


class OutputData:
    investment_fund = ""
    security_type = ""
    security_category = ""
    security_name = 0
    coupon_rate = 0
    face_value = 0
    duration = 0
    weight = 0
    issue_type = ""
    issue_country = ""
    date = ""


def read_xlsfile(xls_filename,xls_date):
    readbook = xlrd.open_workbook(xls_filename)
    sheet = readbook.sheet_by_name("Sheet1")  # 名字的方式
    nrows = sheet.nrows  # 行
    ncols = sheet.ncols  # 列
    for i in range(1,nrows):
        data = OutputData()
        data.investment_fund = sheet.cell_value(i, 0)
        data.security_type = sheet.cell_value(i, 1)
        data.security_category = sheet.cell_value(i, 2)
        data.security_name = sheet.cell_value(i, 3)
        data.coupon_rate = sheet.cell_value(i, 4)
        data.face_value = sheet.cell_value(i, 5)
        data.duration = sheet.cell_value(i, 6)
        data.weight = sheet.cell_value(i, 7)
        data.issue_type = sheet.cell_value(i, 8)
        data.issue_country = sheet.cell_value(i,9)
        data.date = xls_date
        all_data.append(data)


#xls文件
def wt_xlsfile(data):
    book = xlwt.Workbook(encoding='utf-8', style_compression=0)
    sheet = book.add_sheet('alldata', cell_overwrite_ok=True)


    #先写列头
    for i in range(len(para)):
        sheet.write(0,i,para[i])


    #再写入数据
    for i in range(len(data)):
        sheet.write(i + 1, 0, data[i].investment_fund)
        sheet.write(i + 1, 1, data[i].security_type)
        sheet.write(i + 1, 2, data[i].security_category)
        sheet.write(i + 1, 3, data[i].security_name)
        sheet.write(i + 1, 4, data[i].coupon_rate)
        sheet.write(i + 1, 5, data[i].face_value)
        sheet.write(i + 1, 6, data[i].duration)
        sheet.write(i + 1, 7, data[i].weight)
        sheet.write(i + 1, 8, data[i].issue_type)
        sheet.write(i + 1, 9, data[i].issue_country)
        sheet.write(i + 1, 10, data[i].date)
    book.save(r'alldata.xls')


#xlsx文件
def wt_xlsxfile(data):
    # 在内存中创建一个workbook对象，而且会至少创建一个 worksheet
    wb = Workbook()
    # 获取当前活跃的worksheet,默认就是第一个worksheet
    ws = wb.active
    # 先写列头
    for i in range(len(para)):
        ws.cell(row=1, column=i+1).value = para[i]

        # 再写入数据
    for i in range(len(data)):
        ws.cell(row = i + 1, column = 1).value =  data[i].investment_fund
        ws.cell(row=i + 1, column=2).value = data[i].security_type
        ws.cell(row=i + 1, column=3).value = data[i].security_category
        ws.cell(row=i + 1, column=4).value = data[i].security_name
        ws.cell(row=i + 1, column=5).value = data[i].coupon_rate
        ws.cell(row=i + 1, column=6).value = data[i].face_value
        ws.cell(row=i + 1, column=7).value = data[i].duration
        ws.cell(row=i + 1, column=8).value = data[i].weight
        ws.cell(row=i + 1, column=9).value = data[i].issue_type
        ws.cell(row=i + 1, column=10).value = data[i].issue_country
        ws.cell(row=i + 1, column=11).value = data[i].date

    wb.save(filename="alldata.xlsx")


for i in range(MAX):
    file_name = "Dailybondportfolio_" + str(date_start) + ".xlsx"
    readfile_list.append(file_name)
    date_start = date_start + 1
    read_xlsfile(file_name,date_start)

wt_xlsfile(all_data)
wt_xlsxfile(all_data)




